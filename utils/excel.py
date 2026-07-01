"""Utilitários genéricos para leitura e validação de planilhas Excel."""

from __future__ import annotations

import io

import openpyxl
import pandas as pd


def desmesclar_planilha(conteudo: bytes, aba: str) -> io.BytesIO:
    """Preenche células mescladas com o valor da célula superior-esquerda."""
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    if aba not in wb.sheetnames:
        raise ValueError(f"Aba '{aba}' não encontrada. Abas disponíveis: {wb.sheetnames}")

    ws = wb[aba]
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        valor = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row, col).value = valor

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def ler_aba(buffer: io.BytesIO, aba: str, linha_cabecalho: int) -> pd.DataFrame:
    """Lê uma aba Excel usando linha_cabecalho (1-based)."""
    buffer.seek(0)
    return pd.read_excel(buffer, sheet_name=aba, header=linha_cabecalho - 1)


def validar(
    df: pd.DataFrame,
    colunas_obrigatorias: list[str],
    valores_status: set[str] | None = None,
    coluna_status: str = "status",
) -> None:
    """Valida contrato mínimo da planilha; levanta ValueError com mensagem clara."""
    faltando = [c for c in colunas_obrigatorias if c not in df.columns]
    if faltando:
        raise ValueError(f"Colunas obrigatórias ausentes: {faltando}")

    if df.empty:
        raise ValueError("Planilha sem dados após limpeza.")

    vazias = [c for c in colunas_obrigatorias if df[c].isna().all()]
    if vazias:
        raise ValueError(f"Colunas obrigatórias totalmente vazias: {vazias}")

    if valores_status is not None:
        invalidos = set(df[coluna_status].dropna().unique()) - valores_status
        if invalidos:
            raise ValueError(f"Valores de status não reconhecidos: {sorted(invalidos)}")
