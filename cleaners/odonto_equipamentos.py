"""Cleaner da planilha ODONTO — EQUIPAMENTOS PRIORITÁRIOS — serviço Manutenção."""

from __future__ import annotations

import io

import openpyxl
import pandas as pd

from cleaners.equipamentos_prioritarios import _como_texto, _mapear_registro
from utils.excel import validar

COLUNAS_OBRIGATORIAS = ["unidade", "tipo_equipamento", "status"]

STATUS_VALIDOS = {
    "OPERANTE",
    "CONDENADO",
    "PARCIAL OU INOPERANTE",
    "INOPERANTE",
    "DISPONÍVEL",
    "SEM STATUS",
}

_ABAS = {
    "EQUIPOS  CADEIRAS ODONTO": "Cadeira Odontológica",
    "COMPRESSOR ODONTO": "Compressor Odonto",
}

_COLUNAS_FINAIS = [
    "unidade",
    "tipo_equipamento",
    "proprietario",
    "marca",
    "modelo",
    "serie",
    "status",
    "atualizado_em",
    "abertura_os",
    "observacao",
    "sisbens",
    "orgao",
]


def limpar(conteudo: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    registros: list[dict] = []

    for aba, tipo in _ABAS.items():
        if aba not in wb.sheetnames:
            continue
        ws = wb[aba]
        linhas = list(ws.iter_rows(values_only=True))
        if len(linhas) < 2:
            continue
        cabecalho = list(linhas[0])
        for linha in linhas[1:]:
            registro = _mapear_registro(cabecalho, linha, tipo)
            if not registro:
                continue
            dados = dict(zip(cabecalho, linha))
            registro["sisbens"] = _como_texto(dados.get("SISBENS"))
            registro["orgao"] = _como_texto(dados.get("ÓRGÃO") or dados.get("ORGAO"))
            registros.append(registro)
    wb.close()

    df = pd.DataFrame(registros)
    if df.empty:
        raise ValueError("Planilha sem dados após limpeza.")

    for col in _COLUNAS_FINAIS:
        if col not in df.columns:
            df[col] = None

    df = df[_COLUNAS_FINAIS].reset_index(drop=True)
    validar(df, COLUNAS_OBRIGATORIAS, STATUS_VALIDOS)
    return df
