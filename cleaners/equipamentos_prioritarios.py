"""Cleaner da planilha EQUIPAMENTOS PRIORITÁRIOS — serviço Manutenção (várias abas)."""

from __future__ import annotations

import io
import re
from typing import Optional

import openpyxl
import pandas as pd

from utils.excel import validar

COLUNAS_OBRIGATORIAS = ["unidade", "tipo_equipamento", "status"]

STATUS_VALIDOS = {
    "OPERANTE",
    "CONDENADO",
    "PARCIAL OU INOPERANTE",
    "INOPERANTE",
    "EMPRESTADO",
    "NÃO POSSUI",
    "NÃO POSSIU",
    "RETIRADO PARA MANUTENÇÃO",
    "SEM STATUS",
}

_COLUNAS_FINAIS = [
    "unidade",
    "tipo_equipamento",
    "proprietario",
    "marca",
    "modelo",
    "serie",
    "status",
    "atualizado_em",
    "abertura_os",
    "observacao",
    "turno",
    "agenda",
]


def _como_texto(valor) -> Optional[str]:
    if pd.isna(valor):
        return None
    texto = str(valor).strip()
    if texto.endswith(".0") and texto.replace(".", "").isdigit():
        texto = texto[:-2]
    return texto or None


def _normalizar_coluna(nome: str) -> str:
    nome = str(nome).strip().upper()
    nome = re.sub(r"\s+", " ", nome)
    mapa = {
        "1": "UNIDADE",
        "0000": "PROPRIETARIO",
        "E-MAIL": "PROPRIETARIO",
        "Nº DE SÉRIE": "SERIE",
        "N° DE SÉRIE": "SERIE",
        "ATUALIZADO EM": "ATUALIZADO_EM",
        "ABERTURA DE OS": "ABERTURA_OS",
        "OBSERVAÇÃO": "OBSERVACAO",
    }
    return mapa.get(nome, nome)


def _mapear_registro(cabecalho: list, linha: tuple, tipo_equipamento: str) -> dict | None:
    dados = {_normalizar_coluna(k): v for k, v in zip(cabecalho, linha) if k is not None}
    unidade = dados.get("UNIDADE")
    if not unidade or str(unidade).strip().upper() in {"UNIDADE", "1", "NAN"}:
        return None

    status = dados.get("STATUS")
    if status is not None and str(status).strip():
        status_txt = str(status).strip().upper()
        if status_txt == "NÃO POSSIU":
            status_txt = "NÃO POSSUI"
    else:
        status_txt = "SEM STATUS"

    return {
        "unidade": str(unidade).strip(),
        "tipo_equipamento": tipo_equipamento,
        "proprietario": _como_texto(dados.get("PROPRIETARIO")),
        "marca": _como_texto(dados.get("MARCA")),
        "modelo": _como_texto(dados.get("MODELO")),
        "serie": _como_texto(dados.get("SERIE")),
        "status": status_txt,
        "atualizado_em": pd.to_datetime(dados.get("ATUALIZADO_EM"), errors="coerce"),
        "abertura_os": _como_texto(dados.get("ABERTURA_OS")),
        "observacao": _como_texto(dados.get("OBSERVACAO")),
        "turno": _como_texto(dados.get("TURNO")),
        "agenda": _como_texto(dados.get("AGENDA")),
    }


def limpar(conteudo: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    registros: list[dict] = []

    for aba in wb.sheetnames:
        ws = wb[aba]
        linhas = list(ws.iter_rows(values_only=True))
        if len(linhas) < 2:
            continue
        cabecalho = list(linhas[0])
        for linha in linhas[1:]:
            registro = _mapear_registro(cabecalho, linha, aba.strip())
            if registro:
                registros.append(registro)
    wb.close()

    df = pd.DataFrame(registros)
    if df.empty:
        raise ValueError("Planilha sem dados após limpeza.")

    df = df[_COLUNAS_FINAIS].reset_index(drop=True)
    validar(df, COLUNAS_OBRIGATORIAS, STATUS_VALIDOS)
    return df
