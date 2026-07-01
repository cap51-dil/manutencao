"""Cleaner da planilha ACESSO MAIS SEGURO — serviço Manutenção (abas mensais)."""

from __future__ import annotations

import io
import re
from datetime import datetime

import openpyxl
import pandas as pd

COLUNAS_OBRIGATORIAS = ["unidade", "mes", "data", "status_cor"]

CORES_VALIDAS = {"VERDE", "AMARELO", "LARANJA", "VERMELHO"}


def _parse_data(cabecalho) -> datetime | None:
    if isinstance(cabecalho, datetime):
        return cabecalho
    if cabecalho is None:
        return None
    texto = str(cabecalho).strip()
    if not texto or texto.upper() == "UNIDADE":
        return None
    match = re.search(r"(\d{2}/\d{2}/\d{4})", texto)
    if match:
        return datetime.strptime(match.group(1), "%d/%m/%Y")
    return None


def _ler_aba_mes(ws, mes: str) -> list[dict]:
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return []

    cabecalho = linhas[0]
    datas = [_parse_data(col) for col in cabecalho[1:]]
    registros: list[dict] = []

    for linha in linhas[1:]:
        unidade = linha[0]
        if not unidade or str(unidade).strip().upper() == "UNIDADE":
            continue
        unidade = str(unidade).strip()
        for data, status in zip(datas, linha[1:]):
            if data is None or status is None:
                continue
            status_cor = str(status).strip().upper()
            if status_cor not in CORES_VALIDAS:
                continue
            registros.append(
                {
                    "unidade": unidade,
                    "mes": mes,
                    "data": data,
                    "status_cor": status_cor,
                }
            )
    return registros


def limpar(conteudo: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    registros: list[dict] = []
    for aba in wb.sheetnames:
        registros.extend(_ler_aba_mes(wb[aba], aba))
    wb.close()

    df = pd.DataFrame(registros)
    if df.empty:
        raise ValueError("Planilha sem dados após limpeza.")

    df["data"] = pd.to_datetime(df["data"], errors="coerce")
    df = df.dropna(subset=["data"])
    df = df.sort_values(["data", "unidade"]).reset_index(drop=True)

    invalidos = set(df["status_cor"].unique()) - CORES_VALIDAS
    if invalidos:
        raise ValueError(f"Valores de status não reconhecidos: {sorted(invalidos)}")

    for col in COLUNAS_OBRIGATORIAS:
        if df[col].isna().all():
            raise ValueError(f"Colunas obrigatórias totalmente vazias: [{col}]")

    return df
